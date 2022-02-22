from RPA.Excel.Files import XlsWorkbook, XlsxWorkbook, Files
import pathlib

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

class XlsxWorkbookExtend(XlsxWorkbook):

    def open_with_data(self, path):
        path = path or self.path

        if not path:
            raise ValueError("No path defined for workbook")

        try:
            extension = pathlib.Path(path).suffix
        except TypeError:
            extension = None

        if extension in (".xlsm", ".xltm"):
            self._book = openpyxl.load_workbook(filename=path, keep_vba=True)
        else:
            self._book = openpyxl.load_workbook(filename=path, data_only=True)

        self._extension = extension


class FilesExtend(Files):
    def open_workbook_with_data(self, path):
        """Open an existing Excel workbook.

        :param path: path to Excel file
        """
        if self.workbook:
            self.close_workbook()

        self.workbook = self._load_workbook_with_data(path)
        self.logger.info("Opened workbook: %s", self.workbook)
        return self.workbook

    def _load_workbook_with_data(self, path):
        # pylint: disable=broad-except
        path = pathlib.Path(path).resolve(strict=True)

        try:
            book = XlsxWorkbookExtend(path=path)
            book.open_with_data(path)
            return book
        except InvalidFileException as exc:
            self.logger.debug(exc)  # Unsupported extension, silently try xlrd
        except Exception as exc:
            self.logger.info(
                "Failed to open as Office Open XML (.xlsx) format: %s", exc
            )

        try:
            book = XlsWorkbook(path)
            book.open()
            return book
        except Exception as exc:
            self.logger.info("Failed to open as Excel Binary Format (.xls): %s", exc)

        raise ValueError(
            f"Failed to open Excel file ({path}), "
            "verify that the path and extension are correct"
        )
